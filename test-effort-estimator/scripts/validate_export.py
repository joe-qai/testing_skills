#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
验证导出的 Excel 文件
"""

import openpyxl
from collections import Counter


def validate_excel(file_path):
    """
    验证 Excel 文件内容

    Args:
        file_path: Excel 文件路径
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 统计总行数
    total_rows = ws.max_row - 1  # 减去表头
    print(f"总测试用例数: {total_rows}")

    # 统计优先级分布
    priorities = []
    items = Counter()
    points = Counter()

    for row in range(2, ws.max_row + 1):
        priority = ws.cell(row=row, column=3).value
        item = ws.cell(row=row, column=1).value
        point = ws.cell(row=row, column=2).value

        if priority:
            priorities.append(priority)
        if item:
            items[item] += 1
        if point:
            points[point] += 1

    print(f"\n优先级分布:")
    priority_counts = Counter(priorities)
    for priority in ['P0', 'P1', 'P2', 'P3']:
        count = priority_counts.get(priority, 0)
        print(f"  {priority}: {count}")

    print(f"\nITEM 分布 (共 {len(items)} 个):")
    for item, count in items.most_common():
        print(f"  {item}: {count}")

    print(f"\nPOINT 分布 (共 {len(points)} 个):")
    for point, count in points.most_common():
        print(f"  {point}: {count}")

    # 显示前 5 个测试用例
    print(f"\n前 5 个测试用例示例:")
    for row in range(2, min(7, ws.max_row + 1)):
        item = ws.cell(row=row, column=1).value or ""
        point = ws.cell(row=row, column=2).value or ""
        priority = ws.cell(row=row, column=3).value or ""
        title = ws.cell(row=row, column=4).value or ""
        print(f"  {row-1}. [{priority}] {title}")
        print(f"     ITEM: {item}")
        print(f"     POINT: {point}")


if __name__ == '__main__':
    validate_excel(r'C:\pythonworkspace\solo\test-case\export.xlsx')