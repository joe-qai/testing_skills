#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
将 all_cases.md 中的测试用例导出为 Excel 格式
"""

import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_markdown_test_cases(file_path):
    """
    解析 Markdown 文件中的测试用例

    Args:
        file_path: Markdown 文件路径

    Returns:
        包含所有测试用例的列表
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    cases = []
    current_item = ""
    current_point = ""
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # 跳过空行
        if not line:
            i += 1
            continue

        # 检测一级标题 (#) - 跳过
        if line.startswith('# ') and not line.startswith('##'):
            i += 1
            continue

        # 检测二级标题 (##) - ITEM（不包含优先级的标题）
        if line.startswith('## ') and not re.match(r'^##\s+\[P[0-3]\]', line):
            current_item = line[3:].strip()
            current_point = ""
            i += 1
            continue

        # 检测三级标题 (###) - POINT
        if line.startswith('### '):
            current_point = line[4:].strip()
            i += 1
            continue

        # 检测测试用例标题 (## [优先级] 标题)
        test_case_match = re.match(r'^##\s+\[(P[0-3])\]\s+(.+)$', line)
        if test_case_match:
            priority = test_case_match.group(1)
            title = test_case_match.group(2)

            # 解析测试用例的详细信息
            test_type = ""
            precondition = ""
            steps = ""
            expected_result = ""

            # 读取接下来的行来获取详细信息
            j = i + 1
            while j < len(lines):
                detail_line = lines[j].strip()

                # 遇到下一个测试用例或标题时停止
                if detail_line.startswith('##'):
                    break
                if detail_line.startswith('###'):
                    break

                # 解析详细信息
                if detail_line.startswith('[测试类型]'):
                    test_type = detail_line[6:].strip()
                elif detail_line.startswith('[前置条件]'):
                    precondition = detail_line[6:].strip()
                elif detail_line.startswith('[测试步骤]'):
                    steps = detail_line[6:].strip()
                elif detail_line.startswith('[预期结果]'):
                    expected_result = detail_line[6:].strip()
                    # 预期结果可能有多行，继续读取
                    k = j + 1
                    while k < len(lines):
                        next_line = lines[k].strip()
                        if next_line.startswith('##') or next_line.startswith('###') or next_line.startswith('['):
                            break
                        if next_line:  # 如果不是空行，追加到预期结果
                            expected_result += ' ' + next_line
                        k += 1
                    j = k - 1  # 调整 j 的位置

                j += 1

            # 创建测试用例字典
            case = {
                'item': current_item,
                'point': current_point,
                'priority': priority,
                'title': title,
                'test_type': test_type,
                'precondition': precondition,
                'steps': steps,
                'expected_result': expected_result
            }

            cases.append(case)
            i = j
        else:
            i += 1

    return cases


def export_to_excel(cases, output_path):
    """
    将测试用例导出到 Excel 文件

    Args:
        cases: 测试用例列表
        output_path: 输出 Excel 文件路径
    """
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 设置列宽
    ws.column_dimensions['A'].width = 25  # ITEM
    ws.column_dimensions['B'].width = 30  # POINT
    ws.column_dimensions['C'].width = 10  # 优先级
    ws.column_dimensions['D'].width = 45  # 用例标题
    ws.column_dimensions['E'].width = 12  # 测试类型
    ws.column_dimensions['F'].width = 50  # 前置条件
    ws.column_dimensions['G'].width = 60  # 测试步骤
    ws.column_dimensions['H'].width = 60  # 预期结果

    # 设置表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置标题行
    headers = ['ITEM', 'POINT', '优先级', '用例标题', '测试类型', '前置条件', '测试步骤', '预期结果']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 填充数据
    for idx, case in enumerate(cases, 2):
        # 设置单元格值
        ws.cell(row=idx, column=1, value=case['item'])
        ws.cell(row=idx, column=2, value=case['point'])
        ws.cell(row=idx, column=3, value=case['priority'])
        ws.cell(row=idx, column=4, value=case['title'])
        ws.cell(row=idx, column=5, value=case['test_type'])
        ws.cell(row=idx, column=6, value=case['precondition'])
        ws.cell(row=idx, column=7, value=case['steps'])
        ws.cell(row=idx, column=8, value=case['expected_result'])

        # 设置单元格样式
        for col in range(1, 9):
            cell = ws.cell(row=idx, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = thin_border

        # 设置优先级背景色
        priority_cell = ws.cell(row=idx, column=3)
        if case['priority'] == 'P0':
            priority_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif case['priority'] == 'P1':
            priority_cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        elif case['priority'] == 'P2':
            priority_cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        elif case['priority'] == 'P3':
            priority_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

        # 设置行高
        ws.row_dimensions[idx].height = 120

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存文件
    wb.save(output_path)
    print(f"✓ 已成功导出 {len(cases)} 个测试用例到: {output_path}")


def main():
    input_file = r'C:\pythonworkspace\solo\test-case\all_cases.md'
    output_file = r'C:\pythonworkspace\solo\test-case\export.xlsx'

    print("开始解析测试用例...")
    cases = parse_markdown_test_cases(input_file)
    print(f"✓ 共解析到 {len(cases)} 个测试用例")

    print("\n开始导出到 Excel...")
    export_to_excel(cases, output_file)

    print("\n完成！")
    print(f"输出文件: {output_file}")


if __name__ == '__main__':
    main()